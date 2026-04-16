"""
Servicio para generar Excel de Nota de Entrega
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.utils import timezone


def generar_excel_nota_entrega_bytes(nota_id):
    """
    Genera un Excel de la nota de entrega con:
    - Primera hoja: Resumen similar al PDF
    - Segunda hoja: Detalle de items
    
    Args:
        nota_id: ID de la NotaEntrega
        
    Returns:
        bytes: Contenido del Excel
    """
    from reportes.models import NotaEntrega
    
    nota = NotaEntrega.objects.select_related('cliente').prefetch_related('items', 'items__stock', 'items__stock_proveedor', 'items__stock_proveedor__proveedor').get(id=nota_id)
    
    # Calcular totales
    items = nota.items.all()
    subtotal = sum(float(item.precio_unitario_dolares or 0) * item.cantidad for item in items)
    
    # Crear workbook
    wb = Workbook()
    
    # ===== HOJA 1: RESUMEN (como el PDF) =====
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    
    # Estilos
    font_titulo = Font(name='Arial', size=20, bold=True, color='43A29E')
    font_negrita = Font(name='Arial', size=11, bold=True)
    font_normal = Font(name='Arial', size=11)
    font_normal_chico = Font(name='Arial', size=10)
    font_empresa = Font(name='Arial', size=14, bold=True, color='43A29E')
    
    alineacion_centro = Alignment(horizontal='center', vertical='center')
    alineacion_izq = Alignment(horizontal='left', vertical='center')
    alineacion_der = Alignment(horizontal='right', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    fill_azul = PatternFill(start_color='43A29E', end_color='43A29E', fill_type='solid')
    fill_gris = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
    fill_verde = PatternFill(start_color='E6F2F1', end_color='E6F2F1', fill_type='solid')
    
    # Fila 1: Header con logo y banner (como PDF)
    # Logo a la izquierda, banner a la derecha
    ws_resumen.row_dimensions[1].height = 60
    
    # Columna A: Logo (30% del ancho)
    ws_resumen.column_dimensions['A'].width = 15
    ws_resumen['A1'] = "SINCRO"
    ws_resumen['A1'].font = Font(name='Arial', size=24, bold=True, color='43A29E')
    ws_resumen['A1'].alignment = Alignment(horizontal='left', vertical='center')
    
    # Columnas B-G: Banner/Empresa
    ws_resumen.merge_cells('B1:G1')
    ws_resumen['B1'] = "SISTEMA INTEGRADO DE CONSTRUCCIÓN"
    ws_resumen['B1'].font = font_empresa
    ws_resumen['B1'].alignment = Alignment(horizontal='right', vertical='center')
    ws_resumen['B1'].fill = fill_verde
    
    # Columnas para el resto
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws_resumen.column_dimensions[col].width = 18
    
    # Fila 2: Línea divisoria
    ws_resumen.row_dimensions[2].height = 5
    ws_resumen.merge_cells('A2:G2')
    ws_resumen['A2'].border = Border(bottom=Side(style='medium', color='43A29E'))
    
    # Fila 3: Título
    ws_resumen.merge_cells('A3:G3')
    celda_titulo = ws_resumen['A3']
    celda_titulo.value = "NOTA DE ENTREGA"
    celda_titulo.font = font_titulo
    celda_titulo.alignment = alineacion_centro
    ws_resumen.row_dimensions[3].height = 30
    
    # Fila 5: Información de la nota
    ws_resumen['A5'] = "Número:"
    ws_resumen['A5'].font = font_negrita
    ws_resumen['B5'] = nota.numero
    
    ws_resumen['A6'] = "Fecha:"
    ws_resumen['A6'].font = font_negrita
    ws_resumen['B6'] = nota.fecha_creacion.strftime('%d/%m/%Y') if nota.fecha_creacion else "-"
    
    # Fila 8: Cliente
    ws_resumen['A8'] = "Cliente:"
    ws_resumen['A8'].font = font_negrita
    if nota.cliente:
        ws_resumen['B8'] = nota.cliente.nombre
    else:
        ws_resumen['B8'] = nota.cliente_nombre or "—"
    
    if nota.cliente_cedula:
        ws_resumen['A9'] = "Cédula/RIF:"
        ws_resumen['A9'].font = font_negrita
        ws_resumen['B9'] = nota.cliente_cedula
    
    # Fila 11: Encabezados de tabla
    headers = ['#', 'Código', 'Descripción', 'Cantidad', 'Precio Unit.', 'Total']
    for col, header in enumerate(headers, 1):
        celda = ws_resumen.cell(row=11, column=col)
        celda.value = header
        celda.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        celda.fill = fill_azul
        celda.alignment = alineacion_centro
        celda.border = thin_border
    
    # Fila 12+: Items
    row_num = 12
    for idx, item in enumerate(items, 1):
        ws_resumen.cell(row=row_num, column=1, value=idx).border = thin_border
        ws_resumen.cell(row=row_num, column=2, value=item.stock.codigo).border = thin_border
        ws_resumen.cell(row=row_num, column=3, value=item.stock.descripcion).border = thin_border
        ws_resumen.cell(row=row_num, column=4, value=item.cantidad).border = thin_border
        ws_resumen.cell(row=row_num, column=4).alignment = alineacion_der
        ws_resumen.cell(row=row_num, column=5, value=f"${item.precio_unitario_dolares}").border = thin_border
        ws_resumen.cell(row=row_num, column=5).alignment = alineacion_der
        total_item = float(item.precio_unitario_dolares or 0) * item.cantidad
        ws_resumen.cell(row=row_num, column=6, value=f"${total_item:.2f}").border = thin_border
        ws_resumen.cell(row=row_num, column=6).alignment = alineacion_der
        row_num += 1
    
    # Fila de totales
    row_num += 1
    ws_resumen.merge_cells(f'A{row_num}:E{row_num}')
    celda_total = ws_resumen[f'A{row_num}']
    celda_total.value = "TOTAL:"
    celda_total.font = font_negrita
    celda_total.alignment = alineacion_der
    celda_total.fill = fill_gris
    
    ws_resumen[f'F{row_num}'] = f"${subtotal:.2f}"
    ws_resumen[f'F{row_num}'].font = Font(name='Arial', size=11, bold=True, color='43A29E')
    ws_resumen[f'F{row_num}'].alignment = alineacion_der
    ws_resumen[f'F{row_num}'].border = Border(top=Side(style='double'), bottom=Side(style='double'))
    
    # Observaciones
    if nota.observaciones:
        row_num += 2
        ws_resumen[f'A{row_num}'] = "Observaciones:"
        ws_resumen[f'A{row_num}'].font = font_negrita
        ws_resumen.merge_cells(f'A{row_num}:F{row_num}')
        row_num += 1
        ws_resumen.merge_cells(f'A{row_num}:F{row_num}')
        ws_resumen[f'A{row_num}'] = nota.observaciones
        ws_resumen[f'A{row_num}'].font = font_normal_chico
    
    # Pie de página
    row_num += 3
    ws_resumen.merge_cells(f'A{row_num}:F{row_num}')
    ws_resumen[f'A{row_num}'] = f"Documento generado el {timezone.now().strftime('%d/%m/%Y %H:%M')} - SINCRO"
    ws_resumen[f'A{row_num}'].font = Font(name='Arial', size=9, color='999999')
    ws_resumen[f'A{row_num}'].alignment = alineacion_centro
    
    # Ajustar anchos de columna
    ws_resumen.column_dimensions['A'].width = 8
    ws_resumen.column_dimensions['B'].width = 15
    ws_resumen.column_dimensions['C'].width = 40
    ws_resumen.column_dimensions['D'].width = 12
    ws_resumen.column_dimensions['E'].width = 15
    ws_resumen.column_dimensions['F'].width = 15
    
    # ===== HOJA 2: DETALLE =====
    ws_detalle = wb.create_sheet("Detalle")
    
    # Encabezados
    headers_detalle = ['#', 'Código', 'Descripción', 'Proveedor', 'Costo', 'Cantidad', 'Precio Unit.', 'Total']
    for col, header in enumerate(headers_detalle, 1):
        celda = ws_detalle.cell(row=1, column=col)
        celda.value = header
        celda.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        celda.fill = fill_azul
        celda.alignment = alineacion_centro
        celda.border = thin_border
    
    # Datos
    row_num = 2
    for idx, item in enumerate(items, 1):
        ws_detalle.cell(row=row_num, column=1, value=idx).border = thin_border
        ws_detalle.cell(row=row_num, column=2, value=item.stock.codigo).border = thin_border
        ws_detalle.cell(row=row_num, column=3, value=item.stock.descripcion).border = thin_border
        ws_detalle.cell(row=row_num, column=4, value=item.stock_proveedor.proveedor.name).border = thin_border
        ws_detalle.cell(row=row_num, column=5, value=item.cual_costo).border = thin_border
        ws_detalle.cell(row=row_num, column=6, value=item.cantidad).border = thin_border
        ws_detalle.cell(row=row_num, column=6).alignment = alineacion_der
        ws_detalle.cell(row=row_num, column=7, value=float(item.precio_unitario_dolares or 0)).border = thin_border
        ws_detalle.cell(row=row_num, column=7).alignment = alineacion_der
        total_item = float(item.precio_unitario_dolares or 0) * item.cantidad
        ws_detalle.cell(row=row_num, column=8, value=total_item).border = thin_border
        ws_detalle.cell(row=row_num, column=8).alignment = alineacion_der
        row_num += 1
    
    # Totales
    row_num += 1
    ws_detalle.merge_cells(f'A{row_num}:G{row_num}')
    ws_detalle[f'A{row_num}'] = "TOTAL:"
    ws_detalle[f'A{row_num}'].font = font_negrita
    ws_detalle[f'A{row_num}'].alignment = alineacion_der
    ws_detalle[f'H{row_num}'] = subtotal
    ws_detalle[f'H{row_num}'].font = font_negrita
    ws_detalle[f'H{row_num}'].number_format = '$#,##0.00'
    
    # Ajustar anchos
    ws_detalle.column_dimensions['A'].width = 8
    ws_detalle.column_dimensions['B'].width = 15
    ws_detalle.column_dimensions['C'].width = 35
    ws_detalle.column_dimensions['D'].width = 25
    ws_detalle.column_dimensions['E'].width = 12
    ws_detalle.column_dimensions['F'].width = 12
    ws_detalle.column_dimensions['G'].width = 15
    ws_detalle.column_dimensions['H'].width = 15
    
    # Guardar a bytes
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()